import requests as web
import bs4
import openpyxl as excel
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Alignment

# B1などのrange指定文字列に変換する
def cell(col,row):
    col_ar = ['B','C','D','E','F']
    return col_ar[col] + str(row)

# 列番号を列を示す文字に変換する
def col(col):
    col_ar = ['B','C','D','E','F']
    return col_ar[col]

# 新規ワークブックオブジェクトを生成する
wb = excel.Workbook()
# アクティブシートを得る
sheet = wb.active
# シート名を変更する
sheet.title = "報道の自由度ランキング"
# B2セルにタイトルを書く。フォントサイズを24にして、センタリングする
sheet['B2'] = "報道の自由度ランキング"
sheet['B2'].font = Font(size=24)
sheet['B2'].alignment = Alignment(wrap_text=False,  # 折り返し改行
                                  horizontal='center',  # 水平位置
                                  vertical='center'  # 上下位置
        )
# セルを結合する
sheet.merge_cells('B2:F2')
# WEBページにアクセスして、HTMLを取得する
res = web.get('http://ecodb.net/ranking/pfi.html')
if(res.status_code == web.codes.ok):
    # パースしてDOMを取得する
    soup = bs4.BeautifulSoup(res.text,"html.parser")
    # ranking_tableクラスのDIVの内側にあるthタグをすべて取得する
    th = soup.select('div[class=ranking_table] th')
    # thタグをヘダーとして書き込み、背景色とボーダーをつける
    for h in range(len(th)):
        sheet[cell(h,4)] = th[h].getText()
        sheet.row_dimensions[4].height = 20
        sheet.column_dimensions[col(h)].width = 17
        sheet[cell(h,4)].fill = PatternFill(patternType='solid', fgColor='E0FFFF00')
        sheet[cell(h,4)].border = Border(outline=True,
                                         left=Side(style='thin', color='FF000000'),
                                         right=Side(style='thin', color='FF000000'),
                                         top=Side(style='thin', color='FF000000'),
                                         bottom=Side(style='thin', color='FF000000')
        )
        sheet[cell(h,4)].alignment = Alignment(wrap_text=False,  # 折り返し改行
                                       horizontal='general',  # 水平位置
                                       vertical='center'  # 上下位置
        )
    # ranking_tableクラスのDIVの内側にあるtdタグをすべて取得する
    td = soup.select('div[class=ranking_table] td')
    for i in range(len(td)):
        # 1行5項目なので５で割ったり、５の余りをとったりしている
        c = i % 5
        r = (i//5)+5
        # セルに値をセットする
        sheet[cell(c,r)] = td[i].getText()
        # セルの高さを20にする
        if(c == 0):
            sheet.row_dimensions[r].height = 20
        # 罫線を引く
        sheet[cell(c,r)].border = Border(outline=True,
                                                  left=Side(style='thin', color='FF000000'),
                                                  right=Side(style='thin', color='FF000000'),
                                                  top=Side(style='thin', color='FF000000'),
                                                  bottom=Side(style='thin', color='FF000000')
        )
        # 横方向は標準、縦方向は中央にする
        sheet[cell(c,r)].alignment = Alignment(wrap_text=False,  # 折り返し改行
                                       horizontal='general',  # 水平位置
                                       vertical='center'  # 上下位置
        )

else:
    res.raise_for_status()

# EXCELブックを保存する
wb.save('test02.xlsx')
